import openpyxl as xl
from openpyxl.chart import BarChart, Reference #module里import两个class

wb = xl.load_workbook('/Users/ericzhou/Desktop/命令行文件/Texts/Python/Learning Exercises/Excel_Manu/Transactions.xlsx')
sheet = wb['Sheet1'] #后面sheet 要大写

cell1 = sheet['a1']
# 等同于 cell1 = sheet.cell(1,1)
print(cell1.value) #记得显示cell内容时，用.value

# print(sheet.max_row) 最大的row
for row in range(2,sheet.max_row + 1): #从2到最大row，记得加一，不会包括逗号后的值
    cell2 = sheet.cell(row, 3) #每一行第三列(即price)
    number = sheet.cell(row, 4) #每一行第四列(即number)
    total_price = cell2.value * number.value
    print(total_price)
    total_price_cell = sheet.cell(row, 5) #新建的第五行
    total_price_cell.value = total_price #赋值

title = sheet.cell(1,5)
title.value = 'Total' #取标题也要加value！

values = Reference(sheet, #把一个框起来的范围里的值给到reference
 min_row=2, 
 max_row=sheet.max_row, 
 min_col=5, 
 max_col=5)

chart =  BarChart()
chart.add_data(values) #给chart加入reference的values
sheet.add_chart(chart, 'g2')  #在sheet里插入

wb.save('/Users/ericzhou/Desktop/命令行文件/Texts/Python/Learning Exercises/Excel_Manu/transactions2.xlsx') #保存到新文件

